import os
import math
from collections import defaultdict
from openpyxl import Workbook

def calculate_entropy(file_path):
    # Словарь для подсчета количества байтов
    byte_counts = defaultdict(int)

    # Чтение файла побайтно
    with open(file_path, 'rb') as f:
        while (byte := f.read(1)):
            byte_counts[byte[0]] += 1  # Увеличиваем счетчик для данного байта

    # Общее количество байтов в файле
    total_bytes = sum(byte_counts.values())

    # Вычисление вероятностей и энтропии
    entropy = 0.0
    for count in byte_counts.values():
        probability = count / total_bytes
        if probability > 0:  # Избегаем логарифма нуля
            entropy -= probability * math.log2(probability)

    return byte_counts, total_bytes, entropy

def is_printable(byte_value):
    # Проверяем, является ли байт печатным символом (от 32 до 126 в таблице ASCII)
    return 32 <= byte_value <= 126

def save_to_excel(file_name, byte_counts, total_bytes, entropy, ws, start_column):
    # Заполняем заголовки для каждого файла по горизонтали
    ws.cell(row=1, column=start_column, value=f"File: {file_name}")
    ws.cell(row=2, column=start_column, value="Byte")
    ws.cell(row=2, column=start_column + 1, value="Symbol")
    ws.cell(row=2, column=start_column + 2, value="Count")
    ws.cell(row=2, column=start_column + 3, value="Probability")

    # Записываем данные по байтам начиная с 3 строки
    row = 3
    for byte, count in sorted(byte_counts.items()):
        symbol = chr(byte) if is_printable(byte) else '.'
        probability = count / total_bytes
        ws.cell(row=row, column=start_column, value=byte)
        ws.cell(row=row, column=start_column + 1, value=symbol)
        ws.cell(row=row, column=start_column + 2, value=count)
        ws.cell(row=row, column=start_column + 3, value=probability)
        row += 1

    # Добавляем итоговые данные
    ws.cell(row=row + 1, column=start_column, value="Total Bytes")
    ws.cell(row=row + 1, column=start_column + 1, value=total_bytes)
    ws.cell(row=row + 2, column=start_column, value="Entropy")
    ws.cell(row=row + 2, column=start_column + 1, value=entropy)

def main():
    # Создаем новый Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Byte Frequency"

    file_types = ['tif', 'jpg', 'gif', 'bmp']
    start_column = 1  # Начальная колонка для каждого файла

    for file_type in file_types:
        file_path = f"tree.{file_type}"  # Укажите путь к вашему файлу
        if os.path.isfile(file_path):
            byte_counts, total_bytes, entropy = calculate_entropy(file_path)
            print(f"File: {file_path}")
            print(f"Total bytes: {total_bytes}")
            print(f"Entropy: {entropy:.4f}\n")

            # Сохранение данных в Excel для каждого файла, сдвигая по горизонтали
            save_to_excel(file_path, byte_counts, total_bytes, entropy, ws, start_column)

            # Сдвигаем стартовую колонку для следующего файла (4 столбца на файл)
            start_column += 5
        else:
            print(f"File {file_path} does not exist.\n")

    # Сохраняем книгу Excel
    excel_file = "tree.xlsx"
    wb.save(excel_file)
    print(f"Results saved to {excel_file}")

if __name__ == "__main__":
    main()
